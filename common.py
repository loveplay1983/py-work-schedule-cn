# import openpyxl
# from datetime import date, datetime
# import calendar

# # Base class for rest rules
# class RestRule:
#     def is_resting(self, date):
#         pass

#     def is_night_shift(self, date):
#         return False  # Default to no night shift

# # Rest on weekends (Saturday and Sunday)
# class WeekendRestRule(RestRule):
#     def is_resting(self, date):
#         return date.weekday() >= 5  # 5=Saturday, 6=Sunday

# # Rest on fixed days of the week
# class FixedDaysRestRule(RestRule):
#     def __init__(self, rest_days):
#         self.rest_days = rest_days  # List of weekdays (0=Monday, 6=Sunday)

#     def is_resting(self, date):
#         return date.weekday() in self.rest_days

# # Rest periodically, with optional night shifts
# class PeriodicRestRule(RestRule):
#     def __init__(self, rest_every_n_weeks, rest_days, night_shift_every_n_weeks=None, night_shift_days=None):
#         self.rest_every_n_weeks = rest_every_n_weeks
#         self.rest_days = rest_days  # Weekdays for rest
#         self.night_shift_every_n_weeks = night_shift_every_n_weeks or rest_every_n_weeks
#         self.night_shift_days = night_shift_days or []  # Weekdays for night shifts

#     def is_resting(self, date):
#         week_number = date.isocalendar()[1]
#         if week_number % self.rest_every_n_weeks == 1:
#             return date.weekday() in self.rest_days
#         return False

#     def is_night_shift(self, date):
#         week_number = date.isocalendar()[1]
#         if week_number % self.night_shift_every_n_weeks == 0:
#             return date.weekday() in self.night_shift_days
#         return False

# # Fake coworkers with example rules
# coworkers = {
#     "Alice": WeekendRestRule(),              # Rests on weekends, daytime work otherwise
#     "Bob": FixedDaysRestRule([0, 1]),        # Rests on Monday and Tuesday, daytime work otherwise
#     "Charlie": PeriodicRestRule(2, [2], 2, [4])  # Rests every other week on Wednesday, night shift every other week on Friday
# }

# # Generate the schedule with work, night shift, and rest status
# def generate_schedule(year, month, coworkers):
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = f"{calendar.month_name[month]} {year}"

#     # Write headers
#     ws.cell(row=1, column=1, value="Day")
#     for col, coworker in enumerate(coworkers, start=2):
#         ws.cell(row=1, column=col, value=coworker)

#     # Fill schedule
#     num_days = calendar.monthrange(year, month)[1]
#     for day in range(1, num_days + 1):
#         current_date = date(year, month, day)
#         ws.cell(row=day + 1, column=1, value=day)
#         for col, (coworker, rule) in enumerate(coworkers.items(), start=2):
#             if rule.is_resting(current_date):
#                 ws.cell(row=day + 1, column=col, value=f"{day} 休息")
#             elif rule.is_night_shift(current_date):
#                 ws.cell(row=day + 1, column=col, value=f"{day} 值班")
#             else:
#                 ws.cell(row=day + 1, column=col, value=f"{day} 工作")

#     return wb


import openpyxl
from datetime import date, datetime
import calendar

# Weekday names in Chinese
WEEKDAYS = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

# Base class for rest rules
class RestRule:
    def is_resting(self, date):
        pass

    def is_night_shift(self, date):
        return False  # Default to no night shift

# Rest on weekends (Saturday and Sunday)
class WeekendRestRule(RestRule):
    def is_resting(self, date):
        return date.weekday() >= 5  # 5=Saturday, 6=Sunday

# Rest on fixed days of the week
class FixedDaysRestRule(RestRule):
    def __init__(self, rest_days):
        self.rest_days = rest_days  # List of weekdays (0=Monday, 6=Sunday)

    def is_resting(self, date):
        return date.weekday() in self.rest_days

# Rest periodically, with optional night shifts
class PeriodicRestRule(RestRule):
    def __init__(self, rest_every_n_weeks, rest_days, night_shift_every_n_weeks=None, night_shift_days=None):
        self.rest_every_n_weeks = rest_every_n_weeks
        self.rest_days = rest_days  # Weekdays for rest
        self.night_shift_every_n_weeks = night_shift_every_n_weeks or rest_every_n_weeks
        self.night_shift_days = night_shift_days or []  # Weekdays for night shifts

    def is_resting(self, date):
        week_number = date.isocalendar()[1]
        if week_number % self.rest_every_n_weeks == 1:
            return date.weekday() in self.rest_days
        return False

    def is_night_shift(self, date):
        week_number = date.isocalendar()[1]
        if week_number % self.night_shift_every_n_weeks == 0:
            return date.weekday() in self.night_shift_days
        return False

# Fake coworkers with example rules
coworkers = {
    "Alice": WeekendRestRule(),              # Rests on weekends, daytime work otherwise
    "Bob": FixedDaysRestRule([0, 1]),        # Rests on Monday and Tuesday, daytime work otherwise
    "Charlie": PeriodicRestRule(2, [2], 2, [4])  # Rests every other week on Wednesday, night shift every other week on Friday
}

# Generate the schedule with days as columns and employees as rows
def generate_schedule(year, month, coworkers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month]} {year}"

    # Write headers with days and weekdays
    ws.cell(row=1, column=1, value="Employee")
    num_days = calendar.monthrange(year, month)[1]
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        weekday = WEEKDAYS[current_date.weekday()]
        ws.cell(row=1, column=day + 1, value=f"{day} {weekday}")

    # Fill schedule for each employee
    for row, (coworker, rule) in enumerate(coworkers.items(), start=2):
        ws.cell(row=row, column=1, value=coworker)
        for day in range(1, num_days + 1):
            current_date = date(year, month, day)
            if rule.is_resting(current_date):
                ws.cell(row=row, column=day + 1, value=f"{day} 休息")
            elif rule.is_night_shift(current_date):
                ws.cell(row=row, column=day + 1, value=f"{day} 值班")
            else:
                ws.cell(row=row, column=day + 1, value=f"{day} 工作")

    return wb