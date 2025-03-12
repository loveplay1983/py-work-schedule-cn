import openpyxl
from datetime import date, datetime, timedelta
import calendar
from openpyxl.styles import PatternFill, Font
from itertools import cycle
import random

# Simplified weekday names in Chinese
WEEKDAYS = ["一", "二", "三", "四", "五", "六", "日"]

# Chinese month names
MONTH_NAMES = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]

# Colors
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green for rest days
BLUE_FILL = PatternFill(start_color="7db5e3", end_color="7db5e3", fill_type="solid")    # Blue for weekdays
YELLOW_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # Light yellow for weekends

# Statutory holidays for 2025 (example, expand with real dates)
STATUTORY_HOLIDAYS = [
    date(2025, 1, 1),  # New Year's Day
    date(2025, 5, 1),  # Labor Day
    date(2025, 10, 1), # National Day
    # Add more holidays as needed
]

# Base class for rest rules
class RestRule:
    def is_resting(self, date, rest_days=None):
        pass

    def is_night_shift(self, date):
        return False

    def assign_shift(self, date, schedule, rest_days=None):
        pass

# Director Rule (work Mon-Fri, rest Sat-Sun, adjust for statutory holidays)
class DirectorRule(RestRule):
    def is_resting(self, date, rest_days=None):
        if date.weekday() >= 5:  # Rest on weekends
            return True
        if date in STATUTORY_HOLIDAYS and date.weekday() < 5:
            return False  # Work on weekday holidays
        return False  # Work on regular weekdays

    def assign_shift(self, date, schedule, rest_days=None):
        if self.is_resting(date):
            return "休息"
        return "工作"

# Jiangdong Weekend Rule (宣雄民 and 寿春杰)
class JiangdongWeekendRule(RestRule):
    def __init__(self, start_with_jiangdong):
        self.start_with_jiangdong = start_with_jiangdong
        self.jiangdong_cycle = cycle([True, False] if start_with_jiangdong else [False, True])

    def is_resting(self, date, rest_days=None):
        week_number = date.isocalendar()[1]
        is_jiangdong_week = next(self.jiangdong_cycle)
        if is_jiangdong_week and date.weekday() >= 5:
            return False
        if is_jiangdong_week:  # Rest in prior week (Thu/Fri)
            prior_week = week_number - 1
            if prior_week % 2 == (0 if self.start_with_jiangdong else 1):
                return date.weekday() in [3, 4] and date not in rest_days
        return False

    def assign_shift(self, date, schedule, rest_days=None):
        week_number = date.isocalendar()[1]
        is_jiangdong_week = next(self.jiangdong_cycle)
        if date.weekday() >= 5 and is_jiangdong_week:
            return "值班"
        return "工作"

# Weekend Rotation Rule (3 pairs)
class WeekendRotationRule(RestRule):
    def __init__(self, pair_name, other_in_pair, pairs_cycle):
        self.pair_name = pair_name
        self.other_in_pair = other_in_pair
        self.pairs_cycle = pairs_cycle
        self.is_first = True
        self.last_week = -1

    def is_resting(self, date, rest_days=None):
        week_number = date.isocalendar()[1]
        current_cycle = next(self.pairs_cycle)
        if week_number != self.last_week:
            self.is_first = not self.is_first  # Swap roles weekly
            self.last_week = week_number
        if date.weekday() >= 5 and current_cycle:  # Weekend
            if self.is_first:
                return date.weekday() == 5 and self.pair_name == "陈荣盛"
            else:
                return date.weekday() == 6 and self.pair_name == "楼峰"
        # Rest days after weekend duty
        if self.is_first and self.pair_name == "陈荣盛" and date.weekday() in [3, 4]:
            return True
        if not self.is_first and self.pair_name == "楼峰" and date.weekday() in [1, 2]:
            return True
        return False

    def assign_shift(self, date, schedule, rest_days=None):
        week_number = date.isocalendar()[1]
        if week_number != self.last_week:
            self.is_first = not self.is_first  # Swap roles weekly
            self.last_week = week_number
        current_cycle = next(self.pairs_cycle)
        if date.weekday() >= 5 and current_cycle:
            if self.is_first:
                if self.pair_name == "陈荣盛":
                    return "值班" if date.weekday() == 6 else "工作"
                elif self.pair_name == "楼峰":
                    return "工作" if date.weekday() == 6 else "值班"
            else:
                if self.pair_name == "陈荣盛":
                    return "工作" if date.weekday() == 6 else "值班"
                elif self.pair_name == "楼峰":
                    return "值班" if date.weekday() == 6 else "工作"
        return "工作"

# Main Hospital Duty Rule (10 people)
class MainHospitalDutyRule(RestRule):
    def __init__(self, name, all_names):
        self.name = name
        self.all_names = all_names
        self.last_duty = None
        self.duty_count = 0

    def is_resting(self, date, rest_days=None):
        return False

    def assign_shift(self, date, schedule, rest_days=None):
        current_shift = schedule.get((date, self.name), "工作")
        if current_shift in ["江东班", "开发班", "值班", "内勤", "外勤", "休息"]:
            return current_shift

        if self.last_duty and (date - self.last_duty).days < 4:
            return "工作"

        if date.weekday() == 4:  # Friday
            duty_cycle = cycle(self.all_names)
            for _ in range(self.all_names.index(self.name)):
                next(duty_cycle)
            if next(duty_cycle) == self.name:
                self.last_duty = date
                self.duty_count += 1
                return "值班"
        return "工作"

# Internal/External Duty Rule (7 people)
class InternalExternalRule(RestRule):
    def __init__(self, name, internal_group):
        self.name = name
        self.internal_group = internal_group
        self.internal_days = 0
        self.total_working_days = 0

    def is_resting(self, date, rest_days=None):
        return False

    def assign_shift(self, date, schedule, rest_days=None):
        current_shift = schedule.get((date, self.name), "工作")
        if current_shift in ["江东班", "开发班", "值班", "休息"]:
            return current_shift

        if date.weekday() >= 5:  # Weekend
            return "外勤" if current_shift != "外勤" else current_shift

        self.total_working_days += 1
        target_internal_days = self.total_working_days // len(self.internal_group)  # Approx equal distribution
        if self.internal_days < target_internal_days and random.random() < 0.4 and date not in rest_days:
            self.internal_days += 1
            return "内勤"
        return "外勤"

# Development Duty Rule (章杰, 张家栋)
class DevelopmentDutyRule(RestRule):
    def __init__(self, name, other_name):
        self.name = name
        self.other_name = other_name
        self.days_assigned = 0

    def is_resting(self, date, rest_days=None):
        return False

    def assign_shift(self, date, schedule, rest_days=None):
        current_shift = schedule.get((date, self.name), "工作")
        if current_shift == "江东班" or date in rest_days:
            return current_shift

        if self.days_assigned < 4 and date.weekday() in [2, 3, 4]:
            if schedule.get((date, self.other_name), "工作") != "开发班":
                self.days_assigned += 1
                return "开发班"
        return "工作"

# Jiangdong Duty Rule (7 and 9 people)
class JiangdongDutyRule(RestRule):
    def __init__(self, name, group7, group9):
        self.name = name
        self.group7 = group7
        self.group9 = group9
        self.group7_cycle = cycle(group7)
        self.group9_cycle = cycle(group9)

    def is_resting(self, date, rest_days=None):
        return False

    def assign_shift(self, date, schedule, rest_days=None):
        weekday = date.weekday()
        if weekday in [0, 1] and self.name != "郭向彬":  # Mon, Tue
            for _ in range(self.group9.index(self.name)):
                next(self.group9_cycle)
            if next(self.group9_cycle) == self.name and date not in rest_days:
                return "江东班"
        elif weekday in [2, 3, 4]:  # Wed, Thu, Fri
            for _ in range(self.group7.index(self.name)):
                next(self.group7_cycle)
            if next(self.group7_cycle) == self.name and (weekday < 4 or random.random() < 0.5) and date not in rest_days:
                return "江东班"
        elif weekday >= 5:  # Weekend, 1 person
            for _ in range(self.group7.index(self.name)):
                next(self.group7_cycle)
            if next(self.group7_cycle) == self.name and random.random() < 0.2 and date not in rest_days:
                return "江东班"
        return "工作"

# Groups for rotations
main_hospital_duty_names = ["楼峰", "张捷", "郭向彬", "周艺慧", "王振滨", "袁雷武", "章杰", "陈荣盛", "傅舒娜", "张家栋"]
internal_group = ["周艺慧", "王振滨", "袁雷武", "章杰", "陈荣盛", "傅舒娜", "张家栋"]
jiangdong_group7 = ["楼峰", "张捷", "周艺慧", "王振滨", "袁雷武", "陈荣盛", "张家栋"]
jiangdong_group9 = ["楼峰", "张捷", "郭向彬", "周艺慧", "王振滨", "袁雷武", "章杰", "陈荣盛", "张家栋"]

# Coworkers with specific rules
coworkers = {
    "袁铄慧": DirectorRule(),
    "王力天": DirectorRule(),
    "骆飞": DirectorRule(),
    "宣雄民": JiangdongWeekendRule(start_with_jiangdong=True),
    "寿春杰": JiangdongWeekendRule(start_with_jiangdong=False),
    "楼峰": WeekendRotationRule("楼峰", "陈荣盛", cycle([True, False, False])),
    "张捷": WeekendRotationRule("张捷", "袁雷武", cycle([False, True, False])),
    "袁雷武": WeekendRotationRule("袁雷武", "张捷", cycle([False, True, False])),
    "陈荣盛": WeekendRotationRule("陈荣盛", "楼峰", cycle([True, False, False])),
    "王振滨": WeekendRotationRule("王振滨", "章杰", cycle([False, False, True])),
    "章杰": WeekendRotationRule("章杰", "王振滨", cycle([False, False, True])),
    "郭向彬": MainHospitalDutyRule("郭向彬", main_hospital_duty_names),
    "周艺慧": MainHospitalDutyRule("周艺慧", main_hospital_duty_names),
    "傅舒娜": MainHospitalDutyRule("傅舒娜", main_hospital_duty_names),
    "张家栋": MainHospitalDutyRule("张家栋", main_hospital_duty_names),
}

# Add internal/external and Jiangdong rules
for name in internal_group:
    coworkers[name] = InternalExternalRule(name, internal_group)

for name in jiangdong_group7:
    coworkers[name] = JiangdongDutyRule(name, jiangdong_group7, jiangdong_group9)

# Add development duty for 章杰 and 张家栋
coworkers["章杰"] = DevelopmentDutyRule("章杰", "张家栋")
coworkers["张家栋"] = DevelopmentDutyRule("张家栋", "章杰")

# Generate the schedule with days as columns and employees as rows
def generate_schedule(year, month, coworkers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAMES[month-1]} {year}"

    # Write headers
    # Row 1: Day of month with "天" in column 1
    ws.cell(row=1, column=1, value="天")
    ws.cell(row=1, column=1).font = Font(bold=True)
    num_days = calendar.monthrange(year, month)[1]
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        cell = ws.cell(row=1, column=day + 1, value=day)
        cell.font = Font(bold=True)
        if current_date.weekday() >= 5:  # Saturday or Sunday
            cell.fill = YELLOW_FILL
        else:
            cell.fill = BLUE_FILL

    # Row 2: Weekdays with "星期" in column 1
    ws.cell(row=2, column=1, value="星期")
    ws.cell(row=2, column=1).font = Font(bold=True)
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        weekday = WEEKDAYS[current_date.weekday()]
        cell = ws.cell(row=2, column=day + 1, value=weekday)
        cell.font = Font(bold=True)

    # Initialize schedule dictionary
    schedule = {(date(year, month, day), emp): "工作" for day in range(1, num_days + 1) for emp in coworkers.keys()}
    rest_days = {emp: set() for emp in coworkers.keys()}  # Track rest days per employee

    # Assign shifts in priority order: Weekend > Internal/Jiangdong/Development > External
    # Step 1: Directors and Weekend shifts
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        for employee, rule in coworkers.items():
            if isinstance(rule, (DirectorRule, WeekendRotationRule, JiangdongWeekendRule)):
                shift = rule.assign_shift(current_date, schedule, rest_days.get(employee))
                if shift:
                    schedule[(current_date, employee)] = shift
                    if shift == "休息":
                        rest_days[employee].add(current_date)

    # Step 2: Jiangdong and Development shifts
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        for employee, rule in coworkers.items():
            if isinstance(rule, (JiangdongDutyRule, DevelopmentDutyRule)):
                shift = rule.assign_shift(current_date, schedule, rest_days.get(employee))
                if shift:
                    schedule[(current_date, employee)] = shift
                    if shift == "休息":
                        rest_days[employee].add(current_date)

    # Step 3: Main Hospital Duty
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        for employee, rule in coworkers.items():
            if isinstance(rule, MainHospitalDutyRule):
                shift = rule.assign_shift(current_date, schedule, rest_days.get(employee))
                if shift:
                    schedule[(current_date, employee)] = shift
                    if shift == "休息":
                        rest_days[employee].add(current_date)

    # Step 4: Internal/External Duty
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        for employee, rule in coworkers.items():
            if isinstance(rule, InternalExternalRule):
                shift = rule.assign_shift(current_date, schedule, rest_days.get(employee))
                if shift:
                    schedule[(current_date, employee)] = shift
                    if shift == "休息":
                        rest_days[employee].add(current_date)

    # Step 5: Ensure two rest days per week for non-directors
    first_date = date(year, month, 1)
    last_date = date(year, month, num_days)
    for employee in coworkers.keys():
        if isinstance(coworkers[employee], DirectorRule):
            continue
        # Generate week start dates within the month
        week_starts = []
        current_date = first_date
        while current_date <= last_date:
            # Adjust week start to Monday (weekday 0)
            week_start = current_date - timedelta(days=current_date.weekday())
            if week_start < first_date:
                week_start = first_date
            if week_start <= last_date:
                week_starts.append(week_start)
            current_date += timedelta(days=7)

        for week_start in set(week_starts):
            # Only include dates within the current month
            week_dates = [
                week_start + timedelta(days=i)
                for i in range(7)
                if (week_start + timedelta(days=i)).month == month
                and first_date <= (week_start + timedelta(days=i)) <= last_date
            ]
            if not week_dates:
                continue
            rest_count = sum(1 for d in week_dates if schedule[(d, employee)] == "休息")
            while rest_count < 2 and week_dates:
                available_dates = [d for d in week_dates if schedule[(d, employee)] == "工作" and d not in rest_days[employee]]
                if available_dates:
                    rest_date = random.choice(available_dates)
                    schedule[(rest_date, employee)] = "休息"
                    rest_days[employee].add(rest_date)
                    rest_count += 1
                else:
                    break

    # Fill schedule for each employee (starting from row 3)
    for row, (employee, rule) in enumerate(coworkers.items(), start=3):
        ws.cell(row=row, column=1, value=employee)
        for day in range(1, num_days + 1):
            current_date = date(year, month, day)
            cell = ws.cell(row=row, column=day + 1)
            status = schedule[(current_date, employee)]
            if status == "休息":
                cell.value = "休息"
                cell.fill = GREEN_FILL
            else:
                cell.value = status

    return wb